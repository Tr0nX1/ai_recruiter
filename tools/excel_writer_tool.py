"""Append one candidate row to results.xlsx (28 columns)."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from crewai.tools import BaseTool
from pydantic import BaseModel, Field

from config import EXCEL_PATH, JSON_BACKUP_PATH, OUTPUT_DIR

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None  # type: ignore[misc,assignment]
    load_workbook = None  # type: ignore[misc,assignment]

HEADERS = [
    "Candidate Name",
    "Email",
    "Phone",
    "Location",
    "Role Applied",
    "Overall Score",
    "Skills Match %",
    "Experience Score",
    "Education Score",
    "Recommendation",
    "Seniority Level",
    "Summary",
    "Strengths",
    "Weaknesses",
    "Risks",
    "Opportunities",
    "Skills",
    "Missing Skills",
    "Past Experience",
    "Total Years Exp",
    "Education",
    "Certifications",
    "Interview Questions",
    "File Name",
    "Processed Date",
    "Source",
    "Recruiter Notes",
    "Interview Status",
]


def _row_from_dict(row_data: dict[str, Any]) -> list[Any]:
    """Map snake_case keys from blueprint-style dict to column order."""
    key_map = {
        "Candidate Name": "candidate_name",
        "Email": "email",
        "Phone": "phone",
        "Location": "location",
        "Role Applied": "role_applied",
        "Overall Score": "overall_score",
        "Skills Match %": "skills_match_pct",
        "Experience Score": "experience_score",
        "Education Score": "education_score",
        "Recommendation": "recommendation",
        "Seniority Level": "seniority_level",
        "Summary": "summary",
        "Strengths": "strengths",
        "Weaknesses": "weaknesses",
        "Risks": "risks",
        "Opportunities": "opportunities",
        "Skills": "skills",
        "Missing Skills": "missing_skills",
        "Past Experience": "past_experience",
        "Total Years Exp": "total_years_exp",
        "Education": "education",
        "Certifications": "certifications",
        "Interview Questions": "interview_questions",
        "File Name": "file_name",
        "Processed Date": "processed_date",
        "Source": "source",
        "Recruiter Notes": "recruiter_notes",
        "Interview Status": "interview_status",
    }
    out: list[Any] = []
    for h in HEADERS:
        k = key_map[h]
        out.append(row_data.get(k, ""))
    return out


def append_excel_row(row_data: dict[str, Any]) -> dict[str, Any]:
    """Create workbook if needed; append one row; return row number and path."""
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = Path(EXCEL_PATH)
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        ws.append(HEADERS)
        wb.save(path)

    wb = load_workbook(path)
    ws = wb["Candidates"]
    ws.append(_row_from_dict(row_data))
    row_number = ws.max_row
    wb.save(path)
    return {"status": "success", "row_number": row_number, "excel_path": str(path)}


def append_json_backup(entry: dict[str, Any]) -> None:
    path = Path(JSON_BACKUP_PATH)
    data: list[Any] = []
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(data, list):
                data = []
        except json.JSONDecodeError:
            data = []
    data.append(entry)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


class ExcelRowInput(BaseModel):
    row_data: str = Field(
        description="JSON dict with snake_case keys matching excel columns (see tool description).",
    )


class ExcelWriterTool(BaseTool):
    name: str = "ExcelWriterTool"
    description: str = (
        "Appends one candidate row to the Excel report. "
        "Pass row_data as a JSON object with keys: candidate_name, email, phone, location, "
        "role_applied, overall_score, skills_match_pct, experience_score, education_score, "
        "recommendation, seniority_level, summary, strengths, weaknesses, risks, opportunities, "
        "skills, missing_skills, past_experience, total_years_exp, education, certifications, "
        "interview_questions, file_name, processed_date, source, recruiter_notes, interview_status."
    )
    args_schema: type[BaseModel] = ExcelRowInput

    def _run(self, row_data: str) -> dict[str, Any]:
        payload = json.loads(row_data) if isinstance(row_data, str) else row_data
        return append_excel_row(payload)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()
